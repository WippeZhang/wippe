import re
import openpyxl


# excel_file_path = 'aspath-new.xlsx'
# workbook = openpyxl.load_workbook(excel_file_path)
#
# # 选择第一个工作表
# sheet = workbook.active
#
# # 打印表格内容
# data = []
# for row in sheet.iter_rows():
#     for cell in row:
#         # print(cell.value, end='\t')
#         data.append(cell.value)
#
# for i in data:
# # 使用正则表达式提取括号内的数字
#     numbers = []
#     # print(i)
#     num = str(i).split('[')
#     numbers.append(num)
#     for x in numbers:
#         print(x)





def group_and_extract_from_excel(file_path, sheet_name, column_name):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]

    # 获取指定列的数据
    column_data = sheet[column_name]

    data = {}
    for cell in column_data[0:]:
        value = cell.value
        if value and '[' in str(value):
            num, content = str(value).split('[')
            content = content.rstrip(']')

            if content not in data:
                data[content] = [int(num)]
            else:
                data[content].append(int(num))
        if str(value) in str(value) and '[' not in str(value):
            num = int(value)
            if 'Null' not in data:
                data['Null'] = [num]
            else:
                data['Null'].append(num)


    return data

# 文件路径、工作表名称和列名需要根据实际情况修改
file_path = 'aspath.xlsx'
sheet_name = 'Sheet1'
column_name = 'A'

result = group_and_extract_from_excel(file_path, sheet_name, column_name)
# print(result)



for key, values in result.items():
    jz = key
    data = values  # 将 data 定义在循环外
    classified_data = {}
    for num in data:
       #  print(key + ":" + str(num))

        ln = len(str(num)) - 1
        key = str(num)[:ln]  # 取除了个位数之外的数字作为字典键
        value = num % 10  # 取个位数作为字典值
        if key in classified_data:
            classified_data[key].append(value)
        else:
            classified_data[key] = [value]

    classified_data_list = []


    for key, values in sorted(classified_data.items()):
        classified_data_list.append({key: values})


    for xx in classified_data_list:
        print(str(xx) + '[' + str(jz) + ']')

