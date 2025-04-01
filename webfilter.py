import requests
from openpyxl import load_workbook
import pprint

def check_website(url):
    try:
        # response1 = requests.head(url, timeout=12)
        # print(response1.status_code)
        response = requests.get(url, timeout=15)
        print(response.status_code)
        if response.status_code == 200:
            return True
        else:
            return False
    except requests.exceptions.Timeout as a:
        print(a)
        return False
    except requests.ConnectionError as b:
        print(b)
        return False
#
# url = "http://www.example.com"  # 将此处替换为你想要检查的网站
#
# if check_website(url):
#     print(f"{url} 可以访问")
# else:
#     print(f"{url} 无法访问")

wb = load_workbook('深信服URL_Only.xlsx')
sheet = wb.active  # 或者使用 wb['Sheet1'] 来选择特定的工作表
lstcache = []
# 遍历单元格并输出内容
for row in sheet.iter_rows(values_only=True):
    # print(row[0])
    lstcache.append(row[0])

print(len(lstcache))
lstcache = list(set(lstcache))

lst = []
x = 1
for i in lstcache:
    print(x, '---', i)
    url1 = "https://" + i
    url2 = "https://www." + i

    if check_website(str(url1)):
        print('-----------------------', i, '可以访问', '-----------------------')
        lst.append(url1)
    elif check_website(str(url2)):
        print('-----------------------', i, '可以访问', '-----------------------')
        lst.append(url2)
    else:
        print('-----------------------', i, '无法访问', '-----------------------')

    x += 1

pprint.pprint(lst)
print(len(lst))

