from openpyxl import load_workbook

# 读取 Excel 文件
workbook = load_workbook(filename='tls.xlsx')
sheet = workbook.active

# 创建字典以存储 IP 地址关系
left_to_right = {}
right_to_left = {}
output_pairs = []

# 遍历 Excel 表格并填充字典
for row in sheet.iter_rows(values_only=True):
    left_ip, right_ip = row[0].split()
    if right_ip not in left_to_right:
        left_to_right[right_ip] = set()
    left_to_right[right_ip].add(left_ip)

    if left_ip not in right_to_left:
        right_to_left[left_ip] = set()
    right_to_left[left_ip].add(right_ip)

# 汇总左右 IP 地址并输出
for right_ip, left_ips in left_to_right.items():
    if len(left_ips) > 1:
        output_pairs.extend([(left_ip, right_ip) for left_ip in left_ips])

for left_ip, right_ips in right_to_left.items():
    if len(right_ips) > 1:
        output_pairs.extend([(left_ip, right_ip) for right_ip in right_ips])

# 输出剩余未被汇总的数据
for row in sheet.iter_rows(values_only=True):
    left_ip, right_ip = row[0].split()
    if (left_ip, right_ip) not in output_pairs and (right_ip, left_ip) not in output_pairs:
        output_pairs.append((left_ip, right_ip))

# 输出结果
for left_ip, right_ip in output_pairs:
    print(f"{left_ip} {right_ip}")